import datetime

from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMultiAlternatives
from django.shortcuts import render, redirect
from django.contrib.auth.models import User, Group

import openpyxl
from xlwt import Workbook

from receiving.models import LineRequestFinishedHistory, LineRequestFinished, ICDR, ICDRBackup
from shippers.models import Gateway, GatewayBackup
from .models import *
from qad_ee.models import WoMstr


@login_required()
@user_passes_test(lambda u: u.groups.filter(name='crossdocking').exists())
def prodlineFilter(request):
    firstDayPrevMonth = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)).replace(day=15)

    elements = ProdlineTableHistory.objects
    todayItems = ProdlineTable.objects
    return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems, "elements": elements})


@login_required()
@user_passes_test(lambda u: u.groups.filter(name='admin').exists())
def configurations(request):
    prodlines = Prodlines.objects
    users = User.objects
    return render(request, 'crossdocking/configurationsWhProd.html', {"prodlines": prodlines, 'users': users})


def create(request):
    if request.method == 'POST':
        try:
            nome = Prodlines.objects.get(nome=request.POST['nome'].replace(" ", ""))
        except ObjectDoesNotExist:
            prodline = Prodlines()
            prodline.nome = request.POST['nome'].replace(" ", "")
            prodline.save()

            message = 'Adicionada Prodline ' + prodline.nome
            subject, from_email, to = 'Alteração em Crossdocking - Configurations', 'noreply@visteon.com', [
                'aroque1@visteon.com', 'npires2@visteon.com']
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect('crossdocking:configurations')

    prodlines = Prodlines.objects
    return render(request, 'crossdocking/configurationsWhProd.html',
                  {"prodlines": prodlines, 'erro': 'Prodline already exists'})


def delete(request):
    if request.method == 'POST':
        nome = request.POST['nome2']
        try:
            nome = Prodlines.objects.get(nome=nome)
        except ObjectDoesNotExist:
            prodlines = Prodlines.objects
            return render(request, 'crossdocking/configurationsWhProd.html',
                          {"prodlines": prodlines, 'erro': 'Prodline not exists'})

        message = 'Eliminada Prodline ' + nome.nome
        subject, from_email, to = 'Alteração em Crossdocking - Configurations', 'noreply@visteon.com', [
            'aroque1@visteon.com', 'npires2@visteon.com']
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect('crossdocking:configurations')


def submitAll(request):
    if request.method == 'POST':
        user = request.user
        try:
            todayItems = ProdlineTable.objects.all()
            if (request.POST['select'] == "select"):
                if (request.POST['posicao'] == "receiving"):
                    for element in todayItems:
                        element.allOkReceiving = True
                        element.save()
                    enviaEmail(request, user)
                elif (request.POST['posicao'] == "shipping"):
                    for element in todayItems:
                        element.allOkShipping = True
                        element.save()
                    enviaEmailShipping(request, user)
            elif (request.POST['select'] == "deselect"):
                if (request.POST['posicao'] == "receiving"):
                    for element in todayItems:
                        element.allOkReceiving = False
                        element.save()
                    enviaEmail(request, user)
                elif (request.POST['posicao'] == "shipping"):
                    for element in todayItems:
                        element.allOkShipping = False
                        element.save()
                    enviaEmailShipping(request, user)
        except IndexError:
            todayItems = ProdlineTable.objects
            return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems})
    todayItems = ProdlineTable.objects
    return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems})


def changeCheckbox(request):
    if request.method == 'POST':
        id = request.POST['id']

        if (request.POST['shippingCheck'] == 'true'):
            element = ProdlineTable.objects.get(id=id)
            element.shipping = True
            element.save()
        elif (request.POST['shippingCheck'] == 'false'):
            element = ProdlineTable.objects.get(id=id)
            element.shipping = False
            element.save()
        if (request.POST['receivingCheck'] == 'true'):
            element = ProdlineTable.objects.get(id=id)
            element.receiving = True
            element.save()
        elif (request.POST['receivingCheck'] == 'false'):
            element = ProdlineTable.objects.get(id=id)
            element.receiving = False
            element.save()
        todayItems = ProdlineTable.objects
        return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems})


def updateComentReceiving(request):
    if request.method == 'POST':
        id = request.POST['id']
        comentario = request.POST['comentario']

        element = ProdlineTable.objects.get(id=id)
        element.comentarioReceiving = comentario
        element.save()

        todayItems = ProdlineTable.objects
        return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems})


def updateComentShipping(request):
    if request.method == 'POST':
        id = request.POST['id']
        comentario = request.POST['comentario']

        element = ProdlineTable.objects.get(id=id)
        element.comentarioShipping = comentario
        element.save()

        todayItems = ProdlineTable.objects
        return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems})


def schedulePlanning(request):
    if request.method == 'POST':
        prodlines = Prodlines.objects
        path = request.POST.get('textPath')
        list = []

        try:
            workbook = openpyxl.load_workbook(path)
            try:
                worksheet = workbook['Schedule Data']
                for element in worksheet:
                    for prod in prodlines.all():
                        if (prod.nome == element[0].value):
                            elem = {
                                "line": element[0].value,
                                "site": element[1].value,
                                "dueDate": str(element[2].value),
                                "itemNumber": element[3].value,
                                "description": element[4].value,
                                "toComplete": str(element[5].value),
                            }
                            list.append(elem)
                todayItems = ProdlineTable.objects
                elements = ProdlineTableHistory.objects
                return render(request, 'crossdocking/prodlinesFilter.html',
                              {"dadosDia": todayItems, "lista": list, "elements": elements})
            except:
                try:
                    worksheet = workbook['Sheet1']
                    for element in worksheet:
                        for prod in prodlines.all():
                            if (prod.nome == element[0].value):
                                elem = {
                                    "line": element[0].value,
                                    "site": element[1].value,
                                    "dueDate": str(element[2].value),
                                    "itemNumber": element[3].value,
                                    "description": element[4].value,
                                    "toComplete": str(element[5].value),
                                }
                                list.append(elem)
                    todayItems = ProdlineTable.objects
                    elements = ProdlineTableHistory.objects
                    return render(request, 'crossdocking/prodlinesFilter.html',
                                  {"dadosDia": todayItems, "lista": list, "elements": elements})
                except:
                    todayItems = ProdlineTable.objects
                    elements = ProdlineTableHistory.objects
                    return render(request, 'crossdocking/prodlinesFilter.html',
                                  {"dadosDia": todayItems,
                                   "erro5": "Worksheet name not found. Should be: Schedule Data or Sheet1",
                                   "elements": elements})
        except:
            todayItems = ProdlineTable.objects
            elements = ProdlineTableHistory.objects
            return render(request, 'crossdocking/prodlinesFilter.html',
                          {"dadosDia": todayItems, "erro5": "File not exists", "elements": elements})

    elements = ProdlineTableHistory.objects
    todayItems = ProdlineTable.objects
    return render(request, 'crossdocking/prodlinesFilter.html', {"dadosDia": todayItems, "elements": elements})


def changeUserGroups(request):
    if request.method == 'POST':
        user = User.objects.get(username=request.POST['username'])
        grupo = request.POST['paginas']
        if User.objects.filter(username=request.POST['username'], groups__name='crossdocking'):
            my_group = Group.objects.using('default').get(name='crossdocking')
            my_group.user_set.remove(user)
        if grupo == 'none':
            message = 'User ' + user.username + ' perdeu acesso à página de Crossdocking.'
            subject, from_email, to = 'Alteração em Crossdocking - Configurations', 'noreply@visteon.com', [
                'aroque1@visteon.com', 'npires2@visteon.com']
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect('crossdocking:configurations')
        if grupo == 'crossdocking':
            my_group = Group.objects.using('default').get(name='crossdocking')
            my_group.user_set.add(user)
        message = 'User ' + user.username + ' com páginas acessiveis - ' + grupo.replace("/", " , ")
        subject, from_email, to = 'Alteração em Crossdocking - Configurations', 'noreply@visteon.com', [
            'aroque1@visteon.com', 'npires2@visteon.com']
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect('crossdocking:configurations')


def reportCrossdocking(request):
    if request.method == 'POST':
        texto = request.POST['reportICDR'].replace('\n', '</br>')

        message = '<b>Report criado pelo User: ' + request.user.username + '</b>'
        message += "</br></br>" + texto
        subject, from_email, to = 'Novo report na página Crossdocking', 'noreply@visteon.com', [
            'aroque1@visteon.com', 'npires2@visteon.com']
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect('crossdocking:prodlineFilter')


def enviaEmail(request, user):
    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today()
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime('%A') == 'Saturday':
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime('%A') == 'Sunday':
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    mes = nextDay.strftime('%m')
    if (mes == '01'):
        mes = "Janeiro"
    if (mes == '02'):
        mes = "Fevereiro"
    if (mes == '03'):
        mes = "Março"
    if (mes == '04'):
        mes = "Abril"
    if (mes == '05'):
        mes = "Maio"
    if (mes == '06'):
        mes = "Junho"
    if (mes == '07'):
        mes = "Julho"
    if (mes == '08'):
        mes = "Agosto"
    if (mes == '09'):
        mes = "Setembro"
    if (mes == '10'):
        mes = "Outubro"
    if (mes == '11'):
        mes = "Novembro"
    if (mes == '12'):
        mes = "Dezembro"
    mensagem = ''
    textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/' + actualDay.strftime(
        '%Y') + "/" + mes + " " + actualDay.strftime('%Y') + "/Daily_Schedule_" + actualDay.strftime(
        '%d.%m.%Y') + ".xlsx"

    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

    table = '</br>User: <b>' + user.username + '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'

    valueSubmitReceiving = ''
    valueSubmitShipping = ''

    for elem in elementosDiaSeguinte.all():
        valueReceiving = ''
        valueShipping = ''

        if (elem.comentarioShipping == None):
            elem.comentarioShipping = ' '
        if (elem.comentarioReceiving == None):
            elem.comentarioReceiving = ' '
        if (elem.shipping == True):
            valueShipping = 'X'
        if (elem.receiving == True):
            valueReceiving = 'X'
        if (elem.allOkShipping == True):
            valueSubmitShipping = 'X'
        if (elem.allOkReceiving == True):
            valueSubmitReceiving = 'X'
        table += '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">' + elem.line + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.site + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.due_date + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.item_number + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.description + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.to_complete + \
                 '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueReceiving + \
                 '</td><td style="padding:0 15px 0 15px;">' + str(elem.comentarioReceiving) + \
                 '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueShipping + \
                 '</td><td style="padding:0 15px 0 15px;">' + str(elem.comentarioShipping) + '</td></tr>'

    table += '<tr style="background-color: lightgray"><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitReceiving + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitShipping + '</td><td></td></tr>'
    table += '</tbody></table>'
    table += '</br></br><b>Prodlines</b></br>'
    for prod in prodline.all():
        table += prod.nome + '</br>'

    table += '</b></b></br><b>Ficheiro diário de Schedule em: </b>' + textPath
    subject, from_email, to = 'Wh Production ' + actualDay.strftime('%d-%m-%Y'), 'noreply@visteon.com', [
        'npires2@visteon.com', 'abrandao@visteon.com', 'aroque1@visteon.com', 'sanasta1@visteon.com',
        'rsalgue2@visteon.com', 'nlopes8@visteon.com', 'abilro1@visteon.com', 'jrodri80@visteon.com',
        'evenanc1@visteon.com']
    msg = EmailMultiAlternatives(subject, table, from_email, to)
    msg.attach_alternative(table, "text/html")
    msg.send()


def enviaEmailShipping(request, user):
    dadosQAD = WoMstr.objects.values('wo_due_date', 'wo_part', 'wo_qty_exp_complete', 'wo_qty_comp')

    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today()
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime('%A') == 'Saturday':
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime('%A') == 'Sunday':
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    mes = nextDay.strftime('%m')
    if (mes == '01'):
        mes = "Janeiro"
    if (mes == '02'):
        mes = "Fevereiro"
    if (mes == '03'):
        mes = "Março"
    if (mes == '04'):
        mes = "Abril"
    if (mes == '05'):
        mes = "Maio"
    if (mes == '06'):
        mes = "Junho"
    if (mes == '07'):
        mes = "Julho"
    if (mes == '08'):
        mes = "Agosto"
    if (mes == '09'):
        mes = "Setembro"
    if (mes == '10'):
        mes = "Outubro"
    if (mes == '11'):
        mes = "Novembro"
    if (mes == '12'):
        mes = "Dezembro"
    mensagem = ''
    textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/' + actualDay.strftime(
        '%Y') + "/" + mes + " " + actualDay.strftime('%Y') + "/Daily_Schedule_" + actualDay.strftime(
        '%d.%m.%Y') + ".xlsx"

    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

    table = '</br>User: <b>' + user.username + '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'

    valueSubmitReceiving = ''
    valueSubmitShipping = ''

    for elem in elementosDiaSeguinte.all():
        valueReceiving = ''
        valueShipping = ''

        if (elem.comentarioShipping == None):
            elem.comentarioShipping = ' '
        if (elem.comentarioReceiving == None):
            elem.comentarioReceiving = ' '
        if (elem.shipping == True):
            valueShipping = 'X'
        if (elem.receiving == True):
            valueReceiving = 'X'
        if (elem.allOkShipping == True):
            valueSubmitShipping = 'X'
        if (elem.allOkReceiving == True):
            valueSubmitReceiving = 'X'
        table += '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">' + elem.line + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.site + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.due_date + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.item_number + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.description + \
                 '</td><td style="padding:0 15px 0 15px;">' + elem.to_complete + \
                 '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueReceiving + \
                 '</td><td style="padding:0 15px 0 15px;">' + str(elem.comentarioReceiving) + \
                 '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueShipping + \
                 '</td><td style="padding:0 15px 0 15px;">' + str(elem.comentarioShipping) + '</td></tr>'

    table += '<tr style="background-color: lightgray"><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitReceiving + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitShipping + '</td><td></td></tr>'
    table += '</tbody></table>'

    table += '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Qty to Complete</th><th>Actual Qty Completed</th></tr></thead><tbody>'

    for elem3 in elementosDiaSeguinte.all():
        elem2 = dadosQAD.filter(wo_due_date=elem3.due_date, wo_part=elem3.item_number, wo_site='3515')
        for elem1 in elem2.all():
            if str(elem1['wo_qty_exp_complete'])[:-11] != str(elem1['wo_qty_comp'])[:-11]:
                table += '<tr style="background-color: red">' \
                         '<td style="padding:0 15px 0 15px;">' + elem3.line + \
                         '</td><td style="padding:0 15px 0 15px;">' + elem3.site + \
                         '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_due_date']) + \
                         '</td><td style="padding:0 15px 0 15px;">' + elem1['wo_part'] + \
                         '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_exp_complete'])[:-11] + \
                         '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_comp'])[:-11] + '</td></tr>'
            else:
                table += '<tr style="background-color: whitesmoke">' \
                         '<td style="padding:0 15px 0 15px;">' + elem3.line + \
                         '</td><td style="padding:0 15px 0 15px;">' + elem3.site + \
                         '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_due_date']) + \
                         '</td><td style="padding:0 15px 0 15px;">' + elem1['wo_part'] + \
                         '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_exp_complete'])[:-11] + \
                         '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_comp'])[:-11] + '</td></tr>'

    table += '</tbody></table>'
    table += '</br></br><b>Prodlines</b></br>'
    for prod in prodline.all():
        table += prod.nome + '</br>'

    table += '</b></b></br><b>Ficheiro diário de Schedule em: </b>' + textPath
    subject, from_email, to = 'Wh Production ' + actualDay.strftime('%d-%m-%Y'), 'noreply@visteon.com', [
        'npires2@visteon.com', 'abrandao@visteon.com',
        'aroque1@visteon.com', 'sanasta1@visteon.com',
        'rsalgue2@visteon.com', 'nlopes8@visteon.com',
        'abilro1@visteon.com', 'jrodri80@visteon.com',
        'evenanc1@visteon.com']
    msg = EmailMultiAlternatives(subject, table, from_email, to)
    msg.attach_alternative(table, "text/html")
    msg.send()


def enviarEmailSchedule():
    timeNow = datetime.datetime.now()
    timeEnd = timeNow.replace(hour=8, minute=10)
    timeBefore = timeNow.replace(hour=7, minute=55)
    timeEnd2 = timeNow.replace(hour=16, minute=10)
    timeBefore2 = timeNow.replace(hour=15, minute=55)
    timeEnd3 = timeNow.replace(hour=22, minute=10)
    timeBefore3 = timeNow.replace(hour=21, minute=55)

    dadosQAD = WoMstr.objects.values('wo_due_date', 'wo_part', 'wo_qty_exp_complete', 'wo_qty_comp')

    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today()
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime('%A') == 'Saturday':
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime('%A') == 'Sunday':
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    mes = nextDay.strftime('%m')
    if (mes == '01'):
        mes = "Janeiro"
    if (mes == '02'):
        mes = "Fevereiro"
    if (mes == '03'):
        mes = "Março"
    if (mes == '04'):
        mes = "Abril"
    if (mes == '05'):
        mes = "Maio"
    if (mes == '06'):
        mes = "Junho"
    if (mes == '07'):
        mes = "Julho"
    if (mes == '08'):
        mes = "Agosto"
    if (mes == '09'):
        mes = "Setembro"
    if (mes == '10'):
        mes = "Outubro"
    if (mes == '11'):
        mes = "Novembro"
    if (mes == '12'):
        mes = "Dezembro"
    mensagem = ''
    textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/' + nextDay.strftime(
        '%Y') + "/" + mes + " " + nextDay.strftime('%Y') + "/Daily_Schedule_" + nextDay.strftime(
        '%d.%m.%Y') + ".xlsx"

    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

    table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'

    if (timeNow < timeEnd and timeNow > timeBefore) and (actualDay.strftime('%A') != 'Saturday') and (
            actualDay.strftime('%A') != 'Sunday'):
        table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th></tr></thead><tbody>'
        try:
            workbook = openpyxl.load_workbook(textPath)
            try:
                worksheet = workbook['Schedule Data']
                for element in worksheet:
                    for prod in prodline.all():
                        if (prod.nome == element[0].value):
                            table += '<tr><td style="padding:0 15px 0 15px;">' + element[0].value + \
                                     '</td><td style="padding:0 15px 0 15px;">' + element[1].value + \
                                     '</td><td style="padding:0 15px 0 15px;">' + str(element[2].value) + \
                                     '</td><td style="padding:0 15px 0 15px;">' + element[3].value + \
                                     '</td><td style="padding:0 15px 0 15px;">' + element[4].value + \
                                     '</td><td style="padding:0 15px 0 15px;">' + str(
                                element[5].value) + '</td></tr>'

                table += '</tbody></table>'
                table += '</br></br><b>Prodlines</b></br>'
                for prod in prodline.all():
                    table += prod.nome + '</br>'

                table += '</b></b></br><b>Ficheiro diário de Schedule em: </b>' + textPath

                subject, from_email, to = 'Wh Production ' + nextDay.strftime(
                    '%d-%m-%Y'), 'noreply@visteon.com', ['npires2@visteon.com', 'abrandao@visteon.com',
                                                         'aroque1@visteon.com', 'sanasta1@visteon.com',
                                                         'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                                                         'abilro1@visteon.com', 'jrodri80@visteon.com',
                                                         'evenanc1@visteon.com']
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                msg.send()
            except:
                try:
                    worksheet = workbook['Sheet1']
                    for element in worksheet:
                        for prod in prodline.all():
                            if (prod.nome == element[0].value):
                                table += '<tr><td style="padding:0 15px 0 15px;">' + element[0].value + \
                                         '</td><td style="padding:0 15px 0 15px;">' + element[1].value + \
                                         '</td><td style="padding:0 15px 0 15px;">' + str(element[2].value) + \
                                         '</td><td style="padding:0 15px 0 15px;">' + element[3].value + \
                                         '</td><td style="padding:0 15px 0 15px;">' + element[4].value + \
                                         '</td><td style="padding:0 15px 0 15px;">' + str(
                                    element[5].value) + '</td></tr>'

                    table += '</tbody></table>'
                    table += '</br></br><b>Prodlines</b></br>'
                    for prod in prodline.all():
                        table += prod.nome + '</br>'

                    table += '</b></b></br><b>Ficheiro diário de Schedule em: </b>' + textPath

                    subject, from_email, to = 'Wh Production ' + nextDay.strftime(
                        '%d-%m-%Y'), 'noreply@visteon.com', ['npires2@visteon.com', 'abrandao@visteon.com',
                                                             'aroque1@visteon.com', 'sanasta1@visteon.com',
                                                             'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                                                             'abilro1@visteon.com', 'jrodri80@visteon.com',
                                                             'evenanc1@visteon.com']
                    msg = EmailMultiAlternatives(subject, table, from_email, to)
                    msg.attach_alternative(table, "text/html")
                    msg.send()
                except KeyError:
                    mensagem = "</br>Worksheet name not found. Should be: Schedule Data or Sheet1 </br>" + textPath + ""
                    subject, from_email, to = 'Wh Production ' + nextDay.strftime(
                        '%d-%m-%Y'), 'noreply@visteon.com', ['npires2@visteon.com', 'abrandao@visteon.com',
                                                             'aroque1@visteon.com', 'sanasta1@visteon.com',
                                                             'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                                                             'abilro1@visteon.com', 'jrodri80@visteon.com',
                                                             'evenanc1@visteon.com']
                    msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
                    msg.attach_alternative(mensagem, "text/html")
                    msg.send()
        except FileNotFoundError:
            mensagem = '</br> File not found. </br></br>' + textPath + ''
            subject, from_email, to = 'Wh Production ' + nextDay.strftime('%d-%m-%Y'), 'noreply@visteon.com', [
                'npires2@visteon.com', 'abrandao@visteon.com',
                'aroque1@visteon.com', 'sanasta1@visteon.com',
                'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                'abilro1@visteon.com', 'jrodri80@visteon.com',
                'evenanc1@visteon.com']
            msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
            msg.attach_alternative(mensagem, "text/html")
            msg.send()


    elif (timeNow < timeEnd2 and timeNow > timeBefore2) and (not ProdlineTable.objects.filter(
            allOkReceiving=True).exists()) and (actualDay.strftime('%A') != 'Saturday') and (
            actualDay.strftime('%A') != 'Sunday') and (elementosDiaSeguinte.count() > 0):

        valueSubmitReceiving = ''
        valueSubmitShipping = ''

        textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/' + actualDay.strftime(
            '%Y') + "/" + mes + " " + actualDay.strftime('%Y') + "/Daily_Schedule_" + actualDay.strftime(
            '%d.%m.%Y') + ".xlsx"

        for elem in elementosDiaSeguinte.all():
            valueReceiving = ''
            valueShipping = ''

            if (elem.comentarioShipping == None):
                elem.comentarioShipping = ' '
            if (elem.comentarioReceiving == None):
                elem.comentarioReceiving = ' '
            if (elem.shipping == True):
                valueShipping = 'X'
            if (elem.receiving == True):
                valueReceiving = 'X'
            if (elem.allOkShipping == True):
                valueSubmitShipping = 'X'
            if (elem.allOkReceiving == True):
                valueSubmitReceiving = 'X'
            table += '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">' + elem.line + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.site + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.due_date + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.item_number + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.description + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.to_complete + \
                     '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueReceiving + \
                     '</td><td style="padding:0 15px 0 15px;">' + str(elem.comentarioReceiving) + \
                     '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueShipping + \
                     '</td><td style="padding:0 15px 0 15px;">' + str(
                elem.comentarioShipping) + '</td></tr>'

        table += '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitReceiving + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitShipping + '</td><td></td></tr>'
        table += '</tbody></table>'
        table += '</br></br><b>Prodlines</b></br>'
        for prod in prodline.all():
            table += prod.nome + '</br>'

        table += '</b></b></br><b>Ficheiro diário de Schedule em: </b>' + textPath

        subject, from_email, to = 'Wh Production ' + actualDay.strftime(
            '%d-%m-%Y'), 'noreply@visteon.com', ['npires2@visteon.com', 'abrandao@visteon.com',
                                                 'aroque1@visteon.com', 'sanasta1@visteon.com',
                                                 'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                                                 'abilro1@visteon.com', 'jrodri80@visteon.com',
                                                 'evenanc1@visteon.com']
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        msg.send()

    elif (timeNow < timeEnd3 and timeNow > timeBefore3) and (
            not ProdlineTable.objects.filter(allOkReceiving=True).exists() or not ProdlineTable.objects.filter(
        allOkShipping=True).exists()) and (actualDay.strftime('%A') != 'Saturday') and (
            actualDay.strftime('%A') != 'Sunday') and (elementosDiaSeguinte.count() > 0):

        valueSubmitReceiving = ''
        valueSubmitShipping = ''

        textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/' + actualDay.strftime(
            '%Y') + "/" + mes + " " + actualDay.strftime('%Y') + "/Daily_Schedule_" + actualDay.strftime(
            '%d.%m.%Y') + ".xlsx"

        for elem in elementosDiaSeguinte.all():
            valueReceiving = ''
            valueShipping = ''

            if (elem.comentarioShipping == None):
                elem.comentarioShipping = ' '
            if (elem.comentarioReceiving == None):
                elem.comentarioReceiving = ' '
            if (elem.shipping == True):
                valueShipping = 'X'
            if (elem.receiving == True):
                valueReceiving = 'X'
            if (elem.allOkShipping == True):
                valueSubmitShipping = 'X'
            if (elem.allOkReceiving == True):
                valueSubmitReceiving = 'X'
            table += '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">' + elem.line + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.site + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.due_date + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.item_number + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.description + \
                     '</td><td style="padding:0 15px 0 15px;">' + elem.to_complete + \
                     '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueReceiving + \
                     '</td><td style="padding:0 15px 0 15px;">' + str(elem.comentarioReceiving) + \
                     '</td><td style="text-align: center; padding:0 15px 0 15px;">' + valueShipping + \
                     '</td><td style="padding:0 15px 0 15px;">' + str(
                elem.comentarioShipping) + '</td></tr>'

        table += '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitReceiving + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">' + valueSubmitShipping + '</td><td></td></tr>'
        table += '</tbody></table>'

        table += '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Qty to Complete</th><th>Actual Qty Completed</th></tr></thead><tbody>'

        for elem3 in elementosDiaSeguinte.all():
            elem2 = dadosQAD.filter(wo_due_date=elem3.due_date, wo_part=elem3.item_number)
            for elem1 in elem2.all():
                if str(elem1['wo_qty_exp_complete'])[:-11] != str(elem1['wo_qty_comp'])[:-11]:
                    table += '<tr style="background-color: red">' \
                             '<td style="padding:0 15px 0 15px;">' + elem3.line + \
                             '</td><td style="padding:0 15px 0 15px;">' + elem3.site + \
                             '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_due_date']) + \
                             '</td><td style="padding:0 15px 0 15px;">' + elem1['wo_part'] + \
                             '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_exp_complete'])[:-11] + \
                             '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_comp'])[
                                                                          :-11] + '</td></tr>'
                else:
                    table += '<tr style="background-color: whitesmoke">' \
                             '<td style="padding:0 15px 0 15px;">' + elem3.line + \
                             '</td><td style="padding:0 15px 0 15px;">' + elem3.site + \
                             '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_due_date']) + \
                             '</td><td style="padding:0 15px 0 15px;">' + elem1['wo_part'] + \
                             '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_exp_complete'])[:-11] + \
                             '</td><td style="padding:0 15px 0 15px;">' + str(elem1['wo_qty_comp'])[
                                                                          :-11] + '</td></tr>'

        table += '</tbody></table>'

        table += '</br></br><b>Prodlines</b></br>'
        for prod in prodline.all():
            table += prod.nome + '</br>'

        table += '</b></b></br><b>Ficheiro diário de Schedule em: </b>' + textPath

        subject, from_email, to = 'Wh Production ' + actualDay.strftime(
            '%d-%m-%Y'), 'noreply@visteon.com', ['npires2@visteon.com', 'abrandao@visteon.com',
                                                 'aroque1@visteon.com', 'sanasta1@visteon.com',
                                                 'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                                                 'abilro1@visteon.com', 'jrodri80@visteon.com',
                                                 'evenanc1@visteon.com']
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        msg.send()


def updateSchedule():
    prodline = Prodlines.objects

    day = datetime.datetime.today()

    if day.strftime('%A') != 'Saturday' and day.strftime('%A') != 'Sunday':
        mes = day.strftime('%m')
        if (mes == '01'):
            mes = "Janeiro"
        if (mes == '02'):
            mes = "Fevereiro"
        if (mes == '03'):
            mes = "Março"
        if (mes == '04'):
            mes = "Abril"
        if (mes == '05'):
            mes = "Maio"
        if (mes == '06'):
            mes = "Junho"
        if (mes == '07'):
            mes = "Julho"
        if (mes == '08'):
            mes = "Agosto"
        if (mes == '09'):
            mes = "Setembro"
        if (mes == '10'):
            mes = "Outubro"
        if (mes == '11'):
            mes = "Novembro"
        if (mes == '12'):
            mes = "Dezembro"

        textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/' + day.strftime('%Y') + "/" + mes + " " + day.strftime(
            '%Y') + "/Daily_Schedule_" + day.strftime('%d.%m.%Y') + ".xlsx"

        # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

        # guarda elementos do dia anterior na tabela histórico
        elementosDiaAnterior = ProdlineTable.objects
        for elem in elementosDiaAnterior.all():
            if elem.comentarioShipping == None:
                elem.comentarioShipping = ''
            if elem.comentarioReceiving == None:
                elem.comentarioReceiving = ''
            novoElem = ProdlineTableHistory(
                None,
                elem.day,
                elem.line,
                elem.site,
                elem.due_date,
                elem.item_number,
                elem.description,
                elem.to_complete,
                elem.receiving,
                elem.allOkReceiving,
                elem.comentarioReceiving,
                elem.shipping,
                elem.allOkShipping,
                elem.comentarioShipping
            )
            novoElem.save()

        # limpa a tabela provisória
        ProdlineTable.objects.filter().delete()

        workbook = openpyxl.load_workbook(textPath)
        try:
            worksheet = workbook['Schedule Data']
            for element in worksheet:
                for prod in prodline.all():
                    if (prod.nome == element[0].value):
                        novo = ProdlineTable(
                            None,
                            day.strftime('%d-%m-%Y'),
                            element[0].value,
                            element[1].value,
                            str(element[2].value),
                            element[3].value,
                            element[4].value,
                            str(element[5].value),
                            None,
                            None,
                            None,
                            None,
                            None,
                            None
                        )
                        novo.save()
        except:
            worksheet = workbook['Sheet1']
            for element in worksheet:
                for prod in prodline.all():
                    if (prod.nome == element[0].value):
                        novo = ProdlineTable(
                            None,
                            day.strftime('%d-%m-%Y'),
                            element[0].value,
                            element[1].value,
                            str(element[2].value),
                            element[3].value,
                            element[4].value,
                            str(element[5].value),
                            None,
                            None,
                            None,
                            None,
                            None,
                            None
                        )
                        novo.save()


def updateLineRequestDia1():
    lineRequest = LineRequestFinished.objects

    firstDayPrevMonth = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)).replace(day=1)
    prev2MonthName = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%B')
    prev2MonthYear = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%Y')

    wbLineRequest = Workbook()
    sheetLineRequest = wbLineRequest.add_sheet("Sheet 1")

    row2 = 0
    col2 = 0

    sheetLineRequest.write(row2, col2, "Request Date")
    sheetLineRequest.write(row2, col2 + 1, "Part-Number")
    sheetLineRequest.write(row2, col2 + 2, "Line")
    sheetLineRequest.write(row2, col2 + 3, "Require")
    sheetLineRequest.write(row2, col2 + 4, "Receiver")
    sheetLineRequest.write(row2, col2 + 5, "Request Justification")
    sheetLineRequest.write(row2, col2 + 6, "Comment")

    row2 += 1

    for elem in lineRequest.all():
        datetime2 = datetime.datetime.strptime(elem.horaPedido, '%Y-%m-%d %H:%M:%S')
        if datetime2 < firstDayPrevMonth:
            novoRequest = LineRequestFinishedHistory(
                None,
                elem.horaPedido,
                elem.partNumber,
                elem.linha,
                elem.requisitante,
                elem.receiver,
                elem.justificacao,
                elem.comentario
            )
            novoRequest.save()

            sheetLineRequest.write(row2, col2, elem.horaPedido)
            sheetLineRequest.write(row2, col2 + 1, elem.partNumber)
            sheetLineRequest.write(row2, col2 + 2, elem.linha)
            sheetLineRequest.write(row2, col2 + 3, elem.requisitante)
            sheetLineRequest.write(row2, col2 + 4, elem.receiver)
            sheetLineRequest.write(row2, col2 + 5, elem.justificacao)
            sheetLineRequest.write(row2, col2 + 6, elem.comentario)

            row2 += 1

            # elem.delete()
    wbLineRequest.save(
        "W:\\sharedir\\MP&L\\Warehouse\\PWMS\\History\\Line_Request\\workbook_" + prev2MonthName + prev2MonthYear + ".xls")


def updatePortariaDia1():
    portaria = Gateway.objects

    fifteenDayPrevMonth = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)).replace(day=15)
    prev2MonthName = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%B')
    prev2MonthYear = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%Y')

    wbPortaria = Workbook()
    sheetPortaria = wbPortaria.add_sheet("Sheet 1")

    row = 0
    col = 0

    sheetPortaria.write(row, col, "Data/hora chegada")
    sheetPortaria.write(row, col + 1, "Condutor")
    sheetPortaria.write(row, col + 2, "ID")
    sheetPortaria.write(row, col + 3, "Contacto")
    sheetPortaria.write(row, col + 4, "Empresa")
    sheetPortaria.write(row, col + 5, "1ª Matricula")
    sheetPortaria.write(row, col + 6, "2ª Matricula")
    sheetPortaria.write(row, col + 7, "Carga/Descarga")
    sheetPortaria.write(row, col + 8, "Doca")
    sheetPortaria.write(row, col + 9, "Destino carga")
    sheetPortaria.write(row, col + 10, "Tipo de Viatura")
    sheetPortaria.write(row, col + 11, "Data/Hora entrada")
    sheetPortaria.write(row, col + 12, "Abandono")
    sheetPortaria.write(row, col + 13, "Comentarios entrada")
    sheetPortaria.write(row, col + 14, "Data/Hora saida")
    sheetPortaria.write(row, col + 15, "Comentarios saida")

    row += 1

    for elem in portaria.all():
        datetime2 = datetime.datetime.strptime(elem.dataHoraChegada, '%Y-%m-%d %H:%M')
        if datetime2 < fifteenDayPrevMonth:
            abandono = ''
            if elem.abandono == 'true':
                abandono = 'X'

            eleme = GatewayBackup(
                None,
                elem.dataHoraChegada,
                elem.condutor,
                elem.ident,
                elem.contacto,
                elem.empresa,
                elem.primeiraMatricula,
                elem.segundaMatricula,
                elem.cargaDescarga,
                elem.doca,
                elem.destinoCarga,
                elem.tipoViatura,
                elem.dataHoraEntrada,
                abandono,
                elem.comentEntrada,
                elem.dataHoraSaida,
                elem.comentSaida,
            )
            eleme.save()

            sheetPortaria.write(row, col, elem.dataHoraChegada)
            sheetPortaria.write(row, col + 1, elem.condutor)
            sheetPortaria.write(row, col + 2, elem.ident)
            sheetPortaria.write(row, col + 3, elem.contacto)
            sheetPortaria.write(row, col + 4, elem.empresa)
            sheetPortaria.write(row, col + 5, elem.primeiraMatricula)
            sheetPortaria.write(row, col + 6, elem.segundaMatricula)
            sheetPortaria.write(row, col + 7, elem.cargaDescarga)
            sheetPortaria.write(row, col + 8, elem.doca)
            sheetPortaria.write(row, col + 9, elem.destinoCarga)
            sheetPortaria.write(row, col + 10, elem.tipoViatura)
            sheetPortaria.write(row, col + 11, elem.dataHoraEntrada)
            sheetPortaria.write(row, col + 12, abandono)
            sheetPortaria.write(row, col + 13, elem.comentEntrada)
            sheetPortaria.write(row, col + 14, elem.dataHoraSaida)
            sheetPortaria.write(row, col + 15, elem.comentSaida)
            row += 1

            # elem.delete()
    wbPortaria.save(
        "W:\\sharedir\\MP&L\\Warehouse\\PWMS\\History\\Portaria\\workbook_" + prev2MonthName + prev2MonthYear + ".xls")

def updatePortariaDia15():
    portaria = Gateway.objects

    firstDayCurrMonth = datetime.datetime.today().replace(day=1)
    prev2MonthName = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%B')
    prev2MonthYear = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%Y')

    wbPortaria = Workbook()
    sheetPortaria = wbPortaria.add_sheet("Sheet 1")

    row = 0
    col = 0

    sheetPortaria.write(row, col, "Data/hora chegada")
    sheetPortaria.write(row, col + 1, "Condutor")
    sheetPortaria.write(row, col + 2, "ID")
    sheetPortaria.write(row, col + 3, "Contacto")
    sheetPortaria.write(row, col + 4, "Empresa")
    sheetPortaria.write(row, col + 5, "1ª Matricula")
    sheetPortaria.write(row, col + 6, "2ª Matricula")
    sheetPortaria.write(row, col + 7, "Carga/Descarga")
    sheetPortaria.write(row, col + 8, "Doca")
    sheetPortaria.write(row, col + 9, "Destino carga")
    sheetPortaria.write(row, col + 10, "Tipo de Viatura")
    sheetPortaria.write(row, col + 11, "Data/Hora entrada")
    sheetPortaria.write(row, col + 12, "Abandono")
    sheetPortaria.write(row, col + 13, "Comentarios entrada")
    sheetPortaria.write(row, col + 14, "Data/Hora saida")
    sheetPortaria.write(row, col + 15, "Comentarios saida")

    row += 1

    for elem in portaria.all():
        datetime2 = datetime.datetime.strptime(elem.dataHoraChegada, '%Y-%m-%d %H:%M')
        if datetime2 < firstDayCurrMonth:
            abandono = ''
            if elem.abandono == 'true':
                abandono = 'X'

            eleme = GatewayBackup(
                None,
                elem.dataHoraChegada,
                elem.condutor,
                elem.ident,
                elem.contacto,
                elem.empresa,
                elem.primeiraMatricula,
                elem.segundaMatricula,
                elem.cargaDescarga,
                elem.doca,
                elem.destinoCarga,
                elem.tipoViatura,
                elem.dataHoraEntrada,
                abandono,
                elem.comentEntrada,
                elem.dataHoraSaida,
                elem.comentSaida,
            )
            eleme.save()

            sheetPortaria.write(row, col, elem.dataHoraChegada)
            sheetPortaria.write(row, col + 1, elem.condutor)
            sheetPortaria.write(row, col + 2, elem.ident)
            sheetPortaria.write(row, col + 3, elem.contacto)
            sheetPortaria.write(row, col + 4, elem.empresa)
            sheetPortaria.write(row, col + 5, elem.primeiraMatricula)
            sheetPortaria.write(row, col + 6, elem.segundaMatricula)
            sheetPortaria.write(row, col + 7, elem.cargaDescarga)
            sheetPortaria.write(row, col + 8, elem.doca)
            sheetPortaria.write(row, col + 9, elem.destinoCarga)
            sheetPortaria.write(row, col + 10, elem.tipoViatura)
            sheetPortaria.write(row, col + 11, elem.dataHoraEntrada)
            sheetPortaria.write(row, col + 12, abandono)
            sheetPortaria.write(row, col + 13, elem.comentEntrada)
            sheetPortaria.write(row, col + 14, elem.dataHoraSaida)
            sheetPortaria.write(row, col + 15, elem.comentSaida)
            row += 1

            # elem.delete()
    wbPortaria.save(
        "W:\\sharedir\\MP&L\\Warehouse\\PWMS\\History\\Portaria\\workbook_" + prev2MonthName + prev2MonthYear + ".xls")

def updateProductionDia1():
    prodline = ProdlineTableHistory.objects

    firstDayPrevMonth = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)).replace(day=1)
    prev2MonthName = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%B')
    prev2MonthYear = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%Y')

    wbProduction = Workbook()
    sheetProduction = wbProduction.add_sheet("Sheet 1")

    row = 0
    col = 0

    sheetProduction.write(row, col, "Line")
    sheetProduction.write(row, col + 1, "Site")
    sheetProduction.write(row, col + 2, "Due Date")
    sheetProduction.write(row, col + 3, "Item Number")
    sheetProduction.write(row, col + 4, "Description")
    sheetProduction.write(row, col + 5, "To Complete")
    sheetProduction.write(row, col + 6, "Receiving")
    sheetProduction.write(row, col + 7, "Submit Receiving")
    sheetProduction.write(row, col + 8, "Comentario Receiving")
    sheetProduction.write(row, col + 9, "Shipping")
    sheetProduction.write(row, col + 10, "Submit Shipping")
    sheetProduction.write(row, col + 11, "Comentario Shipping")

    row += 1

    for elem in prodline.all():
        datetime2 = datetime.datetime.strptime(elem.day, '%d-%m-%Y')
        if datetime2 < firstDayPrevMonth:
            eleme = ProdlineHistoryAllTime(
                None,
                elem.day,
                elem.line,
                elem.site,
                elem.due_date,
                elem.item_number,
                elem.description,
                elem.to_complete,
                elem.receiving,
                elem.allOkReceiving,
                elem.comentarioReceiving,
                elem.shipping,
                elem.allOkShipping,
                elem.comentarioShipping
            )
            eleme.save()

            sheetProduction.write(row, col, elem.line)
            sheetProduction.write(row, col + 1, elem.site)
            sheetProduction.write(row, col + 2, elem.due_date)
            sheetProduction.write(row, col + 3, elem.item_number)
            sheetProduction.write(row, col + 4, elem.description)
            sheetProduction.write(row, col + 5, elem.to_complete)
            sheetProduction.write(row, col + 6, elem.receiving)
            sheetProduction.write(row, col + 7, elem.allOkReceiving)
            sheetProduction.write(row, col + 8, elem.comentarioReceiving)
            sheetProduction.write(row, col + 9, elem.shipping)
            sheetProduction.write(row, col + 10, elem.allOkShipping)
            sheetProduction.write(row, col + 11, elem.comentarioShipping)
            row += 1

            # elem.delete()
    wbProduction.save(
        "W:\\sharedir\\MP&L\\Warehouse\\PWMS\\History\\Production\\workbook_" + prev2MonthName + prev2MonthYear + ".xls")


def updateICDRDia1():
    icdr = ICDR.objects

    firstDayPrevMonth = (datetime.datetime.today())
    prev2MonthName = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%B')
    prev2MonthYear = (datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)).strftime('%Y')

    wbICDR = Workbook()
    sheetICDR = wbICDR.add_sheet("Sheet 1")

    row = 0
    col = 0

    sheetICDR.write(row, col, "DATA ABERTURA")
    sheetICDR.write(row, col + 1, "AGEING (DIAS)")
    sheetICDR.write(row, col + 2, "Nº/ANO")
    sheetICDR.write(row, col + 3, "FORNECEDOR")
    sheetICDR.write(row, col + 4, "PARTNUMBER")
    sheetICDR.write(row, col + 5, "QUANTIDADE")
    sheetICDR.write(row, col + 6, "MOTIVO")
    sheetICDR.write(row, col + 7, "ALFANDEGA (SIM/NAO)")
    sheetICDR.write(row, col + 8, "RESPONSAVEL (NOME/DEPARATAMENTO)")
    sheetICDR.write(row, col + 9, "COMENTARIOS (ABERTURA)")
    sheetICDR.write(row, col + 10, "UN COST")
    sheetICDR.write(row, col + 11, "TOTAL COST")
    sheetICDR.write(row, col + 12, "RCT UNP")
    sheetICDR.write(row, col + 13, "CYCLE COUNT")
    sheetICDR.write(row, col + 14, "CONSUMO")
    sheetICDR.write(row, col + 15, "QAD (PO Nº)")
    sheetICDR.write(row, col + 16, "COMENTARIOS")
    sheetICDR.write(row, col + 17, "DATA FECHO")
    sheetICDR.write(row, col + 18, "CYCLE COUNT")
    sheetICDR.write(row, col + 19, "DATA CC")
    sheetICDR.write(row, col + 20, "AUDIT CHECK")

    row += 1

    for elem in icdr.all():
        datetime2 = datetime.datetime.strptime(elem.aberturaICDR, '%Y-%m-%d %H:%M:%S')
        if datetime2 < firstDayPrevMonth:

            elemeBackup = ICDRBackup()
            elemeBackup.aberturaICDR = elem.aberturaICDR
            elemeBackup.ageing = elem.ageing
            elemeBackup.nAno = elem.nAno
            elemeBackup.fornecedor = elem.fornecedor
            elemeBackup.partnumber = elem.partnumber
            elemeBackup.quantidade = elem.quantidade
            elemeBackup.tipo = elem.tipo
            elemeBackup.simNao = elem.simNao
            elemeBackup.responsavel = elem.responsavel
            elemeBackup.departamento = elem.departamento
            elemeBackup.comentarioFecho = elem.comentarioFecho
            elemeBackup.unCost = elem.unCost
            elemeBackup.totalCost = elem.totalCost
            elemeBackup.rctUnpCheck = elem.rctUnpCheck
            elemeBackup.cycleCountCheck = elem.cycleCountCheck
            elemeBackup.consumption = elem.consumption
            elemeBackup.po = elem.po
            elemeBackup.comentarioFechoICDR = elem.comentarioFechoICDR
            elemeBackup.date = elem.date
            elemeBackup.cycleCount = elem.cycleCount
            elemeBackup.dataCycleCount = elem.dataCycleCount
            elemeBackup.auditCheck = elem.auditCheck
            elemeBackup.save()

            sheetICDR.write(row, col, elem.aberturaICDR)
            sheetICDR.write(row, col + 1, elem.ageing)
            sheetICDR.write(row, col + 2, elem.nAno)
            sheetICDR.write(row, col + 3, elem.fornecedor)
            sheetICDR.write(row, col + 4, elem.partnumber)
            sheetICDR.write(row, col + 5, elem.quantidade)
            sheetICDR.write(row, col + 6, elem.tipo)
            sheetICDR.write(row, col + 7, elem.simNao)
            sheetICDR.write(row, col + 8, elem.responsavel+"/"+elem.departamento)
            sheetICDR.write(row, col + 9, elem.comentarioFecho)
            sheetICDR.write(row, col + 10, elem.unCost)
            sheetICDR.write(row, col + 11, elem.totalCost)
            sheetICDR.write(row, col + 12, elem.rctUnpCheck)
            sheetICDR.write(row, col + 13, elem.cycleCountCheck)
            sheetICDR.write(row, col + 14, elem.consumption)
            sheetICDR.write(row, col + 15, elem.po)
            sheetICDR.write(row, col + 16, elem.comentarioFechoICDR)
            sheetICDR.write(row, col + 17, elem.date)
            sheetICDR.write(row, col + 18, elem.cycleCount)
            sheetICDR.write(row, col + 19, elem.dataCycleCount)
            sheetICDR.write(row, col + 20, elem.auditCheck)
            row += 1

            # elem.delete()
    wbICDR.save(
        "W:\\sharedir\\MP&L\\Warehouse\\PWMS\\History\\ICDR\\workbook_" + prev2MonthName + prev2MonthYear + ".xls")
